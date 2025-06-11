"""
Report generation functionality for analytics.

Provides PDF and Excel report generation for spending analytics data.
"""

import io
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.contrib.auth import get_user_model

from apps.analytics.models import SpendingAnalytics

User = get_user_model()


class BaseReportGenerator:
    """Base class for report generators."""

    def __init__(self, user, start_date, end_date):
        """
        Initialize the report generator.

        Args:
            user: User instance to generate report for
            start_date: Start date for the report period
            end_date: End date for the report period

        Raises:
            ValueError: If start_date is after end_date
        """
        if start_date > end_date:
            raise ValueError("Start date must be before or equal to end date")

        self.user = user
        self.start_date = start_date
        self.end_date = end_date
        self.analytics = SpendingAnalytics(user, start_date, end_date)


class ExcelReportGenerator(BaseReportGenerator):
    """Generator for Excel-based spending reports."""

    def generate_spending_report(self) -> bytes:
        """
        Generate a comprehensive spending report in Excel format.

        Returns:
            bytes: Excel file data
        """
        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        # Create worksheets
        self._create_summary_sheet(workbook)
        self._create_category_breakdown_sheet(workbook)
        self._create_daily_trends_sheet(workbook)
        self._create_transactions_sheet(workbook)

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_summary_sheet(self, workbook: Workbook) -> None:
        """Create the summary worksheet."""
        ws = workbook.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Spending Report Summary"
        ws["A1"].font = Font(size=16, bold=True)

        # Summary data
        total_spending = self.analytics.get_total_spending()
        transaction_count = self.analytics.get_transaction_count()
        avg_daily = self.analytics.get_average_daily_spending()
        avg_transaction = self.analytics.get_average_transaction_amount()

        # Headers
        ws["A3"] = "Total Spending"
        ws["A4"] = "Transaction Count"
        ws["A5"] = "Average Daily Spending"
        ws["A6"] = "Average Transaction Amount"

        # Values
        ws["B3"] = total_spending
        ws["B4"] = transaction_count
        ws["B5"] = avg_daily
        ws["B6"] = avg_transaction

        # Formatting
        for row in range(3, 7):
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].alignment = Alignment(horizontal="right")

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_category_breakdown_sheet(self, workbook: Workbook) -> None:
        """Create the category breakdown worksheet."""
        ws = workbook.create_sheet("Category Breakdown")

        # Title
        ws["A1"] = "Category Breakdown"
        ws["A1"].font = Font(size=16, bold=True)

        # Headers
        ws["A3"] = "Category"
        ws["B3"] = "Amount"
        for col in ["A3", "B3"]:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        # Data
        category_breakdown = self.analytics.get_category_breakdown()
        sorted_categories = sorted(
            category_breakdown.items(), key=lambda x: x[1], reverse=True
        )

        row = 4
        for category_name, amount in sorted_categories:
            ws[f"A{row}"] = category_name
            ws[f"B{row}"] = amount
            ws[f"B{row}"].alignment = Alignment(horizontal="right")
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    def _create_daily_trends_sheet(self, workbook: Workbook) -> None:
        """Create the daily trends worksheet."""
        ws = workbook.create_sheet("Daily Trends")

        # Title
        ws["A1"] = "Daily Spending Trends"
        ws["A1"].font = Font(size=16, bold=True)

        # Headers
        ws["A3"] = "Date"
        ws["B3"] = "Amount"
        for col in ["A3", "B3"]:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        # Data
        daily_trends = self.analytics.get_spending_trends("daily")

        row = 4
        for trend in daily_trends:
            ws[f"A{row}"] = trend["date"]
            ws[f"B{row}"] = trend["amount"]
            ws[f"B{row}"].alignment = Alignment(horizontal="right")
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15

    def _create_transactions_sheet(self, workbook: Workbook) -> None:
        """Create the transactions worksheet."""
        ws = workbook.create_sheet("Transactions")

        # Title
        ws["A1"] = "Transaction Details"
        ws["A1"].font = Font(size=16, bold=True)

        # Headers
        headers = ["Date", "Category", "Amount", "Merchant", "Notes"]
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}2"] = header
            ws[f"{col_letter}2"].font = Font(bold=True)
            ws[f"{col_letter}2"].fill = PatternFill(
                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
            )

        # Data
        transactions = (
            self.analytics.get_base_queryset()
            .select_related("category")
            .order_by("-date")
        )

        row = 3
        for transaction in transactions:
            ws[f"A{row}"] = transaction.date
            ws[f"B{row}"] = (
                transaction.category.name if transaction.category else "Uncategorized"
            )
            ws[f"C{row}"] = transaction.amount_index
            ws[f"D{row}"] = transaction.merchant or ""
            ws[f"E{row}"] = transaction.notes or ""

            # Alignment
            ws[f"C{row}"].alignment = Alignment(horizontal="right")
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 30


class PDFReportGenerator(BaseReportGenerator):
    """Generator for PDF-based spending reports."""

    def generate_spending_report(self) -> bytes:
        """
        Generate a comprehensive spending report in PDF format.

        Returns:
            bytes: PDF file data
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue,
        )

        # Title
        story.append(Paragraph("Spending Report", title_style))
        story.append(
            Paragraph(
                f"Period: {self.start_date.strftime('%B %d, %Y')} "
                f"to {self.end_date.strftime('%B %d, %Y')}",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 20))

        # Summary section
        self._add_summary_section(story, styles, heading_style)

        # Category breakdown section
        self._add_category_breakdown_section(story, styles, heading_style)

        # Top transactions section
        self._add_top_transactions_section(story, styles, heading_style)

        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def _add_summary_section(
        self, story: List[Any], styles: Any, heading_style: ParagraphStyle
    ) -> None:
        """Add summary section to the PDF."""
        story.append(Paragraph("Summary", heading_style))

        # Get summary data
        total_spending = self.analytics.get_total_spending()
        transaction_count = self.analytics.get_transaction_count()
        avg_daily = self.analytics.get_average_daily_spending()
        avg_transaction = self.analytics.get_average_transaction_amount()

        # Create summary table
        data = [
            ["Metric", "Value"],
            ["Total Spending", f"${total_spending:,.2f}"],
            ["Transaction Count", str(transaction_count)],
            ["Average Daily Spending", f"${avg_daily:,.2f}"],
            ["Average Transaction Amount", f"${avg_transaction:,.2f}"],
        ]

        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(table)
        story.append(Spacer(1, 20))

    def _add_category_breakdown_section(
        self, story: List[Any], styles: Any, heading_style: ParagraphStyle
    ) -> None:
        """Add category breakdown section to the PDF."""
        story.append(Paragraph("Category Breakdown", heading_style))

        category_breakdown = self.analytics.get_category_breakdown()

        if not category_breakdown:
            story.append(
                Paragraph(
                    "No spending data available for this period.", styles["Normal"]
                )
            )
            story.append(Spacer(1, 20))
            return

        # Sort categories by amount (descending)
        sorted_categories = sorted(
            category_breakdown.items(), key=lambda x: x[1], reverse=True
        )

        # Limit to top 10 categories for readability
        top_categories = sorted_categories[:10]

        data = [["Category", "Amount", "Percentage"]]
        total_spending = self.analytics.get_total_spending()

        for category_name, amount in top_categories:
            percentage = (amount / total_spending * 100) if total_spending > 0 else 0
            data.append([category_name, f"${amount:,.2f}", f"{percentage:.1f}%"])

        table = Table(data, colWidths=[2.5 * inch, 1.5 * inch, 1 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(table)
        story.append(Spacer(1, 20))

    def _add_top_transactions_section(
        self, story: List[Any], styles: Any, heading_style: ParagraphStyle
    ) -> None:
        """Add top transactions section to the PDF."""
        story.append(Paragraph("Largest Transactions", heading_style))

        # Get top 10 transactions by amount
        transactions = (
            self.analytics.get_base_queryset()
            .select_related("category")
            .order_by("-amount_index")[:10]
        )

        if not transactions:
            story.append(
                Paragraph("No transactions found for this period.", styles["Normal"])
            )
            return

        data = [["Date", "Category", "Amount", "Merchant"]]

        for transaction in transactions:
            data.append(
                [
                    transaction.date.strftime("%m/%d/%Y"),
                    transaction.category.name
                    if transaction.category
                    else "Uncategorized",
                    f"${transaction.amount_index:,.2f}",
                    transaction.merchant or "N/A",
                ]
            )

        table = Table(data, colWidths=[1 * inch, 1.5 * inch, 1 * inch, 1.5 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(table)
