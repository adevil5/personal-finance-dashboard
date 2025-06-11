"""
Tests for analytics report generation functionality.
"""

import io
from datetime import date, timedelta
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from django.contrib.auth import get_user_model

from apps.analytics.reports import ExcelReportGenerator, PDFReportGenerator
from apps.expenses.models import Transaction
from tests.factories import CategoryFactory, TransactionFactory, UserFactory

User = get_user_model()


@pytest.mark.django_db
class TestExcelReportGenerator:
    """Test Excel report generation functionality."""

    def test_init_requires_user_and_date_range(self):
        """Test that Excel generator requires user and date range."""
        user = UserFactory()
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        generator = ExcelReportGenerator(user, start_date, end_date)

        assert generator.user == user
        assert generator.start_date == start_date
        assert generator.end_date == end_date

    def test_init_validates_date_range(self):
        """Test that Excel generator validates date range."""
        user = UserFactory()
        start_date = date.today()
        end_date = date.today() - timedelta(days=1)

        with pytest.raises(ValueError, match="Start date must be before"):
            ExcelReportGenerator(user, start_date, end_date)

    def test_generate_spending_report_creates_workbook(self):
        """Test that spending report creates proper Excel workbook."""
        user = UserFactory()
        category = CategoryFactory(user=user, name="Groceries")
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        # Create some test transactions
        TransactionFactory.create_batch(
            3,
            user=user,
            category=category,
            amount=Decimal("50.00"),
            date=date.today() - timedelta(days=5),
            transaction_type=Transaction.EXPENSE,
        )

        generator = ExcelReportGenerator(user, start_date, end_date)
        workbook_data = generator.generate_spending_report()

        # Load the workbook to verify structure
        workbook = load_workbook(io.BytesIO(workbook_data))

        # Check that required worksheets exist
        expected_sheets = [
            "Summary",
            "Category Breakdown",
            "Daily Trends",
            "Transactions",
        ]
        for sheet_name in expected_sheets:
            assert sheet_name in workbook.sheetnames

        # Check summary sheet has data
        summary_sheet = workbook["Summary"]
        assert summary_sheet["A1"].value == "Spending Report Summary"
        assert summary_sheet["A3"].value == "Total Spending"
        assert summary_sheet["B3"].value == Decimal("150.00")

    def test_generate_category_breakdown_sheet(self):
        """Test category breakdown sheet generation."""
        user = UserFactory()
        groceries = CategoryFactory(user=user, name="Groceries")
        dining = CategoryFactory(user=user, name="Dining")
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        # Create transactions in different categories
        TransactionFactory(
            user=user,
            category=groceries,
            amount=Decimal("100.00"),
            date=date.today() - timedelta(days=5),
            transaction_type=Transaction.EXPENSE,
        )
        TransactionFactory(
            user=user,
            category=dining,
            amount=Decimal("75.00"),
            date=date.today() - timedelta(days=3),
            transaction_type=Transaction.EXPENSE,
        )

        generator = ExcelReportGenerator(user, start_date, end_date)
        workbook_data = generator.generate_spending_report()

        workbook = load_workbook(io.BytesIO(workbook_data))
        category_sheet = workbook["Category Breakdown"]

        # Check headers
        assert category_sheet["A1"].value == "Category Breakdown"
        assert category_sheet["A3"].value == "Category"
        assert category_sheet["B3"].value == "Amount"

        # Check data (should be sorted by amount descending)
        assert category_sheet["A4"].value == "Groceries"
        assert category_sheet["B4"].value == Decimal("100.00")
        assert category_sheet["A5"].value == "Dining"
        assert category_sheet["B5"].value == Decimal("75.00")

    def test_generate_transactions_sheet(self):
        """Test transactions sheet generation."""
        user = UserFactory()
        category = CategoryFactory(user=user, name="Groceries")
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        transaction = TransactionFactory(
            user=user,
            category=category,
            amount=Decimal("50.00"),
            date=date.today() - timedelta(days=5),
            transaction_type=Transaction.EXPENSE,
            notes="Test transaction",
            merchant="Test Store",
        )

        generator = ExcelReportGenerator(user, start_date, end_date)
        workbook_data = generator.generate_spending_report()

        workbook = load_workbook(io.BytesIO(workbook_data))
        transactions_sheet = workbook["Transactions"]

        # Check headers
        expected_headers = ["Date", "Category", "Amount", "Merchant", "Notes"]
        for i, header in enumerate(expected_headers, 1):
            assert transactions_sheet.cell(row=2, column=i).value == header

        # Check transaction data (Excel converts date to datetime)
        expected_date = transaction.date
        actual_date = transactions_sheet["A3"].value
        if hasattr(actual_date, "date"):
            actual_date = actual_date.date()
        assert actual_date == expected_date
        assert transactions_sheet["B3"].value == "Groceries"
        assert transactions_sheet["C3"].value == Decimal("50.00")
        assert transactions_sheet["D3"].value == "Test Store"
        assert transactions_sheet["E3"].value == "Test transaction"

    def test_generate_report_with_no_data(self):
        """Test report generation with no transactions."""
        user = UserFactory()
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        generator = ExcelReportGenerator(user, start_date, end_date)
        workbook_data = generator.generate_spending_report()

        workbook = load_workbook(io.BytesIO(workbook_data))
        summary_sheet = workbook["Summary"]

        # Should show zero spending
        assert summary_sheet["B3"].value == Decimal("0.00")
        assert summary_sheet["B4"].value == 0  # Transaction count
        assert summary_sheet["B5"].value == Decimal("0.00")  # Average daily


@pytest.mark.django_db
class TestPDFReportGenerator:
    """Test PDF report generation functionality."""

    def test_init_requires_user_and_date_range(self):
        """Test that PDF generator requires user and date range."""
        user = UserFactory()
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        generator = PDFReportGenerator(user, start_date, end_date)

        assert generator.user == user
        assert generator.start_date == start_date
        assert generator.end_date == end_date

    def test_init_validates_date_range(self):
        """Test that PDF generator validates date range."""
        user = UserFactory()
        start_date = date.today()
        end_date = date.today() - timedelta(days=1)

        with pytest.raises(ValueError, match="Start date must be before"):
            PDFReportGenerator(user, start_date, end_date)

    def test_generate_spending_report_creates_pdf(self):
        """Test that spending report creates PDF."""
        user = UserFactory()
        category = CategoryFactory(user=user, name="Groceries")
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        # Create test transaction
        TransactionFactory(
            user=user,
            category=category,
            amount=Decimal("100.00"),
            date=date.today() - timedelta(days=5),
            transaction_type=Transaction.EXPENSE,
        )

        generator = PDFReportGenerator(user, start_date, end_date)
        pdf_data = generator.generate_spending_report()

        # Should return bytes
        assert isinstance(pdf_data, bytes)
        assert len(pdf_data) > 0

    def test_generate_report_includes_summary_data(self):
        """Test that PDF report includes summary data."""
        user = UserFactory()
        category = CategoryFactory(user=user, name="Groceries")
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        # Create test transactions
        TransactionFactory.create_batch(
            3,
            user=user,
            category=category,
            amount=Decimal("50.00"),
            date=date.today() - timedelta(days=5),
            transaction_type=Transaction.EXPENSE,
        )

        generator = PDFReportGenerator(user, start_date, end_date)

        # Test the analytics data that would be included
        analytics = generator.analytics
        assert analytics.get_total_spending() == Decimal("150.00")
        assert analytics.get_transaction_count() == 3

    def test_generate_report_with_no_data(self):
        """Test PDF generation with no transactions."""
        user = UserFactory()
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        generator = PDFReportGenerator(user, start_date, end_date)
        pdf_data = generator.generate_spending_report()

        # Should still generate PDF even with no data
        assert isinstance(pdf_data, bytes)
        assert len(pdf_data) > 0


@pytest.mark.django_db
class TestReportIntegration:
    """Test report generation integration scenarios."""

    def test_reports_include_only_user_data(self):
        """Test that reports only include data for the specified user."""
        user1 = UserFactory()
        user2 = UserFactory()
        category = CategoryFactory(user=user1, name="Groceries")
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        # Create transactions for both users
        TransactionFactory(
            user=user1,
            category=category,
            amount=Decimal("100.00"),
            date=date.today() - timedelta(days=5),
            transaction_type=Transaction.EXPENSE,
        )
        TransactionFactory(
            user=user2,
            category=CategoryFactory(user=user2, name="Dining"),
            amount=Decimal("50.00"),
            date=date.today() - timedelta(days=3),
            transaction_type=Transaction.EXPENSE,
        )

        # Generate report for user1
        excel_generator = ExcelReportGenerator(user1, start_date, end_date)
        workbook_data = excel_generator.generate_spending_report()

        workbook = load_workbook(io.BytesIO(workbook_data))
        summary_sheet = workbook["Summary"]

        # Should only show user1's spending
        assert summary_sheet["B3"].value == Decimal("100.00")

    def test_reports_respect_date_range(self):
        """Test that reports only include transactions in the date range."""
        user = UserFactory()
        category = CategoryFactory(user=user, name="Groceries")
        start_date = date.today() - timedelta(days=15)
        end_date = date.today() - timedelta(days=5)

        # Transaction inside range
        TransactionFactory(
            user=user,
            category=category,
            amount=Decimal("100.00"),
            date=date.today() - timedelta(days=10),
            transaction_type=Transaction.EXPENSE,
        )

        # Transaction outside range (too old)
        TransactionFactory(
            user=user,
            category=category,
            amount=Decimal("50.00"),
            date=date.today() - timedelta(days=20),
            transaction_type=Transaction.EXPENSE,
        )

        # Transaction outside range (too recent)
        TransactionFactory(
            user=user,
            category=category,
            amount=Decimal("75.00"),
            date=date.today() - timedelta(days=2),
            transaction_type=Transaction.EXPENSE,
        )

        excel_generator = ExcelReportGenerator(user, start_date, end_date)
        workbook_data = excel_generator.generate_spending_report()

        workbook = load_workbook(io.BytesIO(workbook_data))
        summary_sheet = workbook["Summary"]

        # Should only include the transaction in range
        assert summary_sheet["B3"].value == Decimal("100.00")

    def test_reports_handle_large_datasets(self):
        """Test report generation with large number of transactions."""
        user = UserFactory()
        category = CategoryFactory(user=user, name="Groceries")
        start_date = date.today() - timedelta(days=30)
        end_date = date.today()

        # Create 100 transactions
        TransactionFactory.create_batch(
            100,
            user=user,
            category=category,
            amount=Decimal("10.00"),
            date=date.today() - timedelta(days=10),
            transaction_type=Transaction.EXPENSE,
        )

        excel_generator = ExcelReportGenerator(user, start_date, end_date)
        workbook_data = excel_generator.generate_spending_report()

        workbook = load_workbook(io.BytesIO(workbook_data))

        # Check summary
        summary_sheet = workbook["Summary"]
        assert summary_sheet["B3"].value == Decimal("1000.00")  # 100 * 10.00

        # Check transactions sheet has all records
        transactions_sheet = workbook["Transactions"]
        # Headers in row 2, data starts from row 3
        last_row = transactions_sheet.max_row
        assert last_row == 102  # 1 title + 1 header + 100 transactions
